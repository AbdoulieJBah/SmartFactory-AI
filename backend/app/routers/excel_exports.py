from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from database import get_db
from dependencies import get_current_company, get_current_user
from models import (
    Company,
    DowntimeRecord,
    Inventory,
    Product,
    ProductionOrder,
    QualityCheck,
    User,
)

router = APIRouter(prefix="/excel", tags=["Excel Exports"])


def excel_response(workbook: Workbook, filename: str):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/inventory")
def export_inventory_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company: Company = Depends(get_current_company),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    ws.append(["ID", "Product", "Warehouse", "Quantity", "Reserved", "Min Stock", "Max Stock"])

    items = (
        db.query(Inventory)
        .filter(Inventory.company_id == company.id)
        .order_by(Inventory.id.desc())
        .all()
    )

    for item in items:
        ws.append([
            item.id,
            item.product.name if item.product else "Unknown",
            item.warehouse,
            item.quantity,
            item.reserved_quantity,
            item.min_stock,
            item.max_stock,
        ])

    return excel_response(wb, "inventory.xlsx")


@router.get("/production-orders")
def export_production_orders_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company: Company = Depends(get_current_company),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Orders"

    ws.append([
        "Order Number",
        "Product",
        "Work Center",
        "Target Quantity",
        "Produced Quantity",
        "Priority",
        "Status",
    ])

    orders = (
        db.query(ProductionOrder)
        .filter(ProductionOrder.company_id == company.id)
        .order_by(ProductionOrder.id.desc())
        .all()
    )

    for order in orders:
        ws.append([
            order.order_number,
            order.product.name if order.product else "Unknown",
            order.work_center.name if order.work_center else "N/A",
            order.target_quantity,
            order.produced_quantity,
            order.priority,
            order.status,
        ])

    return excel_response(wb, "production_orders.xlsx")


@router.get("/quality-checks")
def export_quality_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company: Company = Depends(get_current_company),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quality Checks"

    ws.append([
        "ID",
        "Production Order ID",
        "Check Type",
        "Result",
        "Inspector",
        "Defects Count",
        "Corrective Action",
        "Notes",
    ])

    checks = (
        db.query(QualityCheck)
        .filter(QualityCheck.company_id == company.id)
        .order_by(QualityCheck.id.desc())
        .all()
    )

    for check in checks:
        ws.append([
            check.id,
            check.production_order_id,
            check.check_type,
            check.result,
            check.inspector_name,
            check.defects_count,
            check.corrective_action,
            check.notes,
        ])

    return excel_response(wb, "quality_checks.xlsx")


@router.get("/oee")
def export_oee_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company: Company = Depends(get_current_company),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "OEE Report"

    planned_minutes = 480

    downtime = (
        db.query(func.sum(DowntimeRecord.duration_minutes))
        .filter(DowntimeRecord.company_id == company.id)
        .scalar()
    ) or 0

    runtime = max(planned_minutes - downtime, 0)

    availability = (runtime / planned_minutes * 100) if planned_minutes > 0 else 0

    total_target = (
        db.query(func.sum(ProductionOrder.target_quantity))
        .filter(ProductionOrder.company_id == company.id)
        .scalar()
    ) or 0

    total_produced = (
        db.query(func.sum(ProductionOrder.produced_quantity))
        .filter(ProductionOrder.company_id == company.id)
        .scalar()
    ) or 0

    performance = (total_produced / total_target * 100) if total_target > 0 else 0

    checks = (
        db.query(QualityCheck)
        .filter(QualityCheck.company_id == company.id)
        .count()
    )

    failed = (
        db.query(QualityCheck)
        .filter(
            QualityCheck.company_id == company.id,
            QualityCheck.result == "Fail",
        )
        .count()
    )

    quality = (((checks - failed) / checks) * 100) if checks > 0 else 100

    oee = (availability * performance * quality) / 10000

    ws.append(["Metric", "Value"])
    ws.append(["Company", company.name])
    ws.append(["Planned Minutes", planned_minutes])
    ws.append(["Runtime Minutes", round(runtime, 2)])
    ws.append(["Downtime Minutes", round(downtime, 2)])
    ws.append(["Availability %", round(availability, 2)])
    ws.append(["Performance %", round(performance, 2)])
    ws.append(["Quality %", round(quality, 2)])
    ws.append(["OEE %", round(oee, 2)])
    ws.append(["Total Target Quantity", round(total_target, 2)])
    ws.append(["Total Produced Quantity", round(total_produced, 2)])
    ws.append(["Total Quality Checks", checks])
    ws.append(["Failed Quality Checks", failed])

    return excel_response(wb, "oee_report.xlsx")